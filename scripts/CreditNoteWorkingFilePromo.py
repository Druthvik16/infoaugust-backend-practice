from openpyxl import load_workbook
import pandas as pd
import requests
import json
import sys
import os
from datetime import date
pd.options.mode.chained_assignment = None

try:
    #def outputSuccessFile(outputFolderName, outputFileName, outputSheetName, df, checkColumn, checkValue):
        #df_Success = df[(df[checkColumn] == checkValue) & (df['CNNumber'] != '0') & (df['CNNumber'] != '')]
    def outputSuccessFile(outputFolderName, outputFileName, outputSheetName, df, BillToCode, CNNumber,PostingDate):
        df_Success = df[(df['BillToCode'] == BillToCode) & (df['CNNumber'] == CNNumber) & (df['PostingDate'] == PostingDate) &
                        (df['CNNumber'] != '0') & (df['CNNumber'] != '')]
        with pd.ExcelWriter(os.path.join(outputFolderName, outputFileName), engine='openpyxl', mode='a') as writer:
            df_Success.to_excel(writer, sheet_name=outputSheetName, index=False)

    def outputFailedFile(outputFolderName, outputFileName, outputSheetName, df, checkColumn, checkList):
        df_Failed = df[(df[checkColumn].isin(checkList)) | (df['CNNumber'] == '0') | (df['CNNumber'] == '')]
        with pd.ExcelWriter(os.path.join(outputFolderName, outputFileName), engine='openpyxl', mode='a') as writer:
            df_Failed.to_excel(writer, sheet_name=outputSheetName, index=False)

    def checkFormat(columnName, dfList):
        for df in dfList:
            if columnName not in df.columns:
                print(json.dumps({"data" :"Format error: File is missing column ["+columnName+"]",  "code": "FILEERROR"}))
                sys.exit()

    def creditNoteNoExists(cnNo, cnDate):
        cnDate = pd.to_datetime(cnDate, format='mixed', dayfirst=True).strftime('%Y-%m-%d')
        filterCreditNoteNumbers = [creditNoteNumber for creditNoteNumber in creditNoteNumbers if creditNoteNumber['creditNoteNumber'] == cnNo and creditNoteNumber['postingDate'] == cnDate]
        if len(filterCreditNoteNumbers) > 0:
            return True
        return False

    def getPartnerLocationRemark(partnerCode):
        filterPartnerLocations = [partnerLocation for partnerLocation in partnerLocations if partnerLocation['code'] == partnerCode]
        if len(filterPartnerLocations) > 0:
            if filterPartnerLocations[0]['customerType'] == 'MBO':
                return ''
            return 'Invalid customer type'
        return 'Partner location not found in database'

    def getClientDocMasterId(cnNo, cnDate):
        filterCreditNoteNumbers = [creditNoteNumber['id'] for creditNoteNumber in creditNoteNumbers if creditNoteNumber['creditNoteNumber'] == cnNo and creditNoteNumber['postingDate'] == cnDate]
        return filterCreditNoteNumbers[0]

    def uploadWorkingFile(fileName, file, creditNoteNumber, clientUUID, partnerLocationCode, postingDate):
        postingDate = pd.to_datetime(postingDate, format='mixed', dayfirst=True).strftime('%Y-%m-%d')

        api_URL = base_URL + 'api/uploadedDoc/uploadWorkingFile'
        payload = {
           "client" : json.dumps({"uuid" : clientUUID}),
           "documentNewFolderPath" : sys.argv[5],
           "creditNoteNumber" : creditNoteNumber,
           "documentAttachment" : json.dumps({"id" : getDocumentAttachmentId()}),
           "clientDocMasterId" : json.dumps({"id" : getClientDocMasterId(creditNoteNumber + '-' + partnerLocationCode, postingDate)}),
           "partnerLocationCode": partnerLocationCode,
           "postingDate": postingDate
        }

        print(json.dumps("payload" + str(payload)))
        
        files=[
              ('uploadFile', (fileName, file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        ]
        response = requests.post(api_URL, data=payload, headers={}, files=files)
        print(json.dumps('Save working file payload-------'))
        print(json.dumps(payload))
        print(json.dumps('Save working file response-------'))
        print(json.dumps(response.text))
        responseJSON = json.loads(response.text)
        if response.status_code == 200:
            return response.status_code
        else:
            return responseJSON.get('message')

    def getRemarkForCreditNote(cnNo, locationCode, postingDate):
        postingDate = pd.to_datetime(postingDate, format='mixed', dayfirst=True).strftime('%Y-%m-%d')      

        api_URL = base_URL + 'api/uploadedDoc/getRemarkForCreditNote/' + cnNo + '/' + locationCode + '/' + postingDate

        print(json.dumps('get remark for credit note url-------'))
        print(json.dumps(api_URL))
        if mode == 'dev':
            return 'No remark'
        response = requests.get(api_URL)
        print(json.dumps('get remark for credit note response-------'))
        print(json.dumps(response.text))
        responseJSON = json.loads(response.text)
        return responseJSON.get("data").get("remark")

    def uploadFailedFile(fileName, file, clientUUID):
        api_URL = base_URL + 'api/uploadedDoc/uploadFailedFile'
        payload = {
           "client" : json.dumps({"uuid" : clientUUID}),
           "documentFailedFolderPath" : sys.argv[6]
        }
        
        files=[
              ('uploadFile', (fileName, file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        ]

        response = requests.post(api_URL, data=payload, headers={}, files=files)
        print(json.dumps('Save failed file payload-------'))
        print(json.dumps(payload))
        print(json.dumps('Save failed file response-------'))
        print(json.dumps(response.text))
        responseJSON = json.loads(response.text)
        return response.status_code

    def deleteIfExists(filePath):
        if os.path.exists(filePath) and mode != 'dev':
            os.remove(filePath)

    def getDocumentAttachmentId():
        documentAttachmentIds = [documentAttachment['id'] for documentAttachment in documentAttachments if documentAttachment['name'] == 'Working File']
        return documentAttachmentIds[0]

    def changeDateFormat(dateFormat, columnName, dfList):
        for df in dfList:
            df[columnName] = pd.to_datetime(df[columnName], format='mixed', dayfirst=True, errors='coerce')
            df[columnName] = df[columnName].dt.strftime(dateFormat)
    
    #Variable declaration
    mode = 'prod'
    # mode = 'dev'
    if mode == 'dev':
        base_URL = ''
        inputFilePath = r"C:\Divyansh\CreditNoteWorkingFile\MBO_Discount_Jan25.xlsx" 
            # PCNW_20-02-24_1.xlsx    r"C:\Divyansh\APProcess\CreditNoteWorkingFile\FOFO_Discount_Aug24.xlsx"
        creditNoteNumberFilePath = r'C:\Divyansh\CreditNoteWorkingFile\billNoCNW_MBO.txt'    #r'C:\Divyansh\APProcess\CreditNoteWorkingFile\billNo.txt'
        partnerLocationsFilePath = r'C:\Divyansh\CreditNoteWorkingFile\partnerLocationCNW_MBO.txt'  #r'C:\Divyansh\APProcess\CreditNoteWorkingFile\partnerLocations.txt'
        documentAttachments = []
        documentCategories = []
        documents = []
        clientUUID = '' #clientUUID 
    else:
        base_URL = sys.argv[8]
        inputFilePath = sys.argv[4]
        creditNoteNumberFilePath = sys.argv[7]
        partnerLocationsFilePath = sys.argv[9]
        documentAttachments = json.loads(sys.argv[1])
        clientUUID = sys.argv[3]

    outputFolderName = 'scripts'

    file = open(creditNoteNumberFilePath, 'r')
    creditNoteNumbers = json.loads(file.read())
    file.close()
    
    file = open(partnerLocationsFilePath, 'r')
    partnerLocations = json.loads(file.read())
    file.close()
    
    isSuccess = False
    df_Summary_Failed_List = []
    requiredSheets = ['Summary', 'Eoss_1',  'Eoss_2',  'Eoss_3', 'Promo_1', 'Promo_2', 'Promo_3']      
    #['CNNumber', 'PostingDate', 'BillToCode']
    #Process start
    print(json.dumps('Process started...'))
    print(json.dumps('Reading files...'))

    wb = load_workbook(inputFilePath, read_only=True)
    for sheetName in requiredSheets:
        if sheetName not in wb.sheetnames:
            print(json.dumps({"data":"Error occured: Sheet " + sheetName + " not found", "code": "FILEERROR"}))
            sys.exit()
    
    inputFileName = os.path.basename(inputFilePath)
    df_Summary = pd.read_excel(inputFilePath, sheet_name='Summary', dtype=str).fillna('')
    #df_Eoss_1
    df_Eoss_1 = pd.read_excel(inputFilePath, sheet_name='Eoss_1', dtype=str)       #'Eoss_2' #aug tech , inodeed, service base, service application as saas
    df_Eoss_2 = pd.read_excel(inputFilePath, sheet_name='Eoss_2', dtype=str)
    df_Eoss_3 = pd.read_excel(inputFilePath, sheet_name='Eoss_3', dtype=str)

    df_Promo_1 = pd.read_excel(inputFilePath, sheet_name='Promo_1', dtype=str)
    df_Promo_2 = pd.read_excel(inputFilePath, sheet_name='Promo_2', dtype=str)
    df_Promo_3 = pd.read_excel(inputFilePath, sheet_name='Promo_3', dtype=str)

    #Format checker
    print(json.dumps('Checking format...'))
    checkFormat("CNNumber", [df_Summary, df_Eoss_1, df_Eoss_2, df_Eoss_3, df_Promo_1, df_Promo_2, df_Promo_3] ) # [df_Summary, df_EOSS, df_Promo_1]
    checkFormat("PostingDate",[df_Summary, df_Eoss_1, df_Eoss_2, df_Eoss_3, df_Promo_1, df_Promo_2, df_Promo_3] )  # [df_Summary, df_EOSS, df_Promo_1]
    checkFormat("BillToCode", [df_Summary, df_Eoss_1, df_Eoss_2, df_Eoss_3, df_Promo_1, df_Promo_2, df_Promo_3] )   # [df_Summary, df_EOSS, df_Promo_1]
    print(json.dumps('Format OK'))
    print(json.dumps('Processing files...'))

    df_Failed_CNNo = df_Summary[(df_Summary['CNNumber'] == '0') | (df_Summary['CNNumber'] == '')]
    df_Failed_CNNo['Remark'] = 'Invalid CN Number'
    df_Summary = df_Summary[(df_Summary['CNNumber'] != '0') & (df_Summary['CNNumber'] != '')]
    
    #Changing date format
    changeDateFormat("%d-%m-%Y", "PostingDate",[df_Summary, df_Eoss_1, df_Eoss_2, df_Eoss_3, df_Promo_1, df_Promo_2, df_Promo_3] ) #   [df_Summary, df_EOSS, df_Promo_1]

    rowNo = 1
    for index, row in df_Summary.iterrows():
        print(json.dumps("Processing credit note working row no. " + str(rowNo) + " of " + str(len(df_Summary))))
        rowNo += 1

        partnerLocationRemark = getPartnerLocationRemark(row["BillToCode"])
        creditNoteRemark = ''
        if not creditNoteNoExists(row["CNNumber"] + '-' + row["BillToCode"], row["PostingDate"]):
            creditNoteRemark = getRemarkForCreditNote(row["CNNumber"], row["BillToCode"], row["PostingDate"])
        
        if partnerLocationRemark != '':
            row["Remark"] = partnerLocationRemark
            df_Summary_Failed_List.append(row.to_dict())
        elif creditNoteRemark != '':
            row["Remark"] = creditNoteRemark
            df_Summary_Failed_List.append(row.to_dict())
        else:
            df_Summary_Success = df_Summary[(df_Summary["CNNumber"] == row["CNNumber"]) & (df_Summary["BillToCode"] == row["BillToCode"]) & (df_Summary["PostingDate"] == row["PostingDate"])]
            
            outputFileName = row["CNNumber"] + '-' + row["BillToCode"] + '_' + row["PostingDate"] + '_' + date.today().strftime("%d-%m-%Y") + '.xlsx'
            df_Summary_Success.to_excel(os.path.join(outputFolderName, outputFileName), sheet_name='Summary', index=False)
            print(json.dumps("Output file generated: " + outputFileName))
            
            #outputSuccessFile(outputFolderName, outputFileName, "Eoss_1", df_Eoss_1, "BillToCode", row["BillToCode"])   #outputSuccessFile(outputFolderName, outputFileName, "Eoss_2", df_Eoss_1, "BillToCode", row["BillToCode"])
            outputSuccessFile(outputFolderName, outputFileName, "Eoss_1", df_Eoss_1, row["BillToCode"], row["CNNumber"], row["PostingDate"])
            print(json.dumps("Output file generated: 1"))

            #outputSuccessFile(outputFolderName, outputFileName, "Eoss_2", df_Eoss_2, "BillToCode", row["BillToCode"])   #outputSuccessFile(outputFolderName, outputFileName, "Eoss_1", df_Eoss_2, "BillToCode", row["BillToCode"])
            outputSuccessFile(outputFolderName, outputFileName, "Eoss_2", df_Eoss_2, row["BillToCode"], row["CNNumber"], row["PostingDate"])
            print(json.dumps("Output file generated: 2"))

            #outputSuccessFile(outputFolderName, outputFileName, "Eoss_3", df_Eoss_3, "BillToCode", row["BillToCode"])
            outputSuccessFile(outputFolderName, outputFileName, "Eoss_3", df_Eoss_3, row["BillToCode"], row["CNNumber"], row["PostingDate"])
            print(json.dumps("Output file generated: 3"))

            #outputSuccessFile(outputFolderName, outputFileName, "Promo_1", df_Promo_1, "BillToCode", row["BillToCode"])
            outputSuccessFile(outputFolderName, outputFileName, "Promo_1", df_Promo_1, row["BillToCode"], row["CNNumber"], row["PostingDate"])
            print(json.dumps("Output file generated: 4"))

            #outputSuccessFile(outputFolderName, outputFileName, "Promo_2", df_Promo_2, "BillToCode", row["BillToCode"])
            outputSuccessFile(outputFolderName, outputFileName, "Promo_2", df_Promo_2, row["BillToCode"], row["CNNumber"], row["PostingDate"])
            print(json.dumps("Output file generated: 5"))

            #outputSuccessFile(outputFolderName, outputFileName, "Promo_3", df_Promo_3, "BillToCode", row["BillToCode"])
            outputSuccessFile(outputFolderName, outputFileName, "Promo_3", df_Promo_3, row["BillToCode"], row["CNNumber"], row["PostingDate"])
            print(json.dumps("Output file generated: 5.1"))

            file = open(os.path.join(outputFolderName, outputFileName),"rb")
            print(json.dumps("Output file generated: 6"))
            print(json.dumps({"msg" : outputFileName, "cred num" : row["CNNumber"], "clientUuid" : clientUUID}))

            if mode != 'dev':
                statusCode = uploadWorkingFile(outputFileName, file, row["CNNumber"], clientUUID, row["BillToCode"], row['PostingDate'])
            else:
                statusCode = 200
            print(json.dumps({"msg" : "Output file generated: 7", "statusCode" : statusCode}))

            if statusCode == 200:
                isSuccess = True
                print(json.dumps("Output file generated: 8"))
            else:
                df_Summary_Failed_List.append(df_Summary_Success)

            file.close()
            print(json.dumps("Output file generated: 9"))

            deleteIfExists(os.path.join(outputFolderName, outputFileName))
            print(json.dumps("Output file generated: 10"))


    print(json.dumps("From " + str(len(df_Summary)) + " credit note working "  + str(len(df_Summary_Failed_List)) + " rows rejected and rest all accepted"))
    if len(df_Summary_Failed_List) > 0:
        outputFileName = os.path.splitext(inputFileName)[0]+'.xlsx'
        #df_Summary_Failed = pd.concat(df_Summary_Failed_List, axis=0, ignore_index=True)
        df_Summary_Failed = pd.DataFrame(df_Summary_Failed_List)
        if(len(df_Failed_CNNo) > 0):
            df_Summary_Failed = pd.concat([df_Summary_Failed, df_Failed_CNNo], axis=0, ignore_index=True)
        
        if 'Remark' not in df_Summary_Failed.columns:
            df_Summary_Failed.insert(len(df_Summary_Failed.columns), 'Remark', '')
        df_Summary_Failed.to_excel(os.path.join(outputFolderName, outputFileName), sheet_name='Summary', index=False)
        print(json.dumps("Output file generated with remark: " + outputFileName))

        outputFailedFile(outputFolderName, outputFileName, "Eoss_1", df_Eoss_1, "BillToCode", df_Summary["BillToCode"])     #outputFailedFile(outputFolderName, outputFileName, "Eoss_2", df_Eoss_1, "BillToCode", df_Summary["BillToCode"])
        print(json.dumps("Output file generated with remark: 1"))

        outputFailedFile(outputFolderName, outputFileName, "Eoss_2", df_Eoss_2, "BillToCode", df_Summary["BillToCode"])     #outputFailedFile(outputFolderName, outputFileName, "Eoss_1", df_Eoss_2, "BillToCode", df_Summary["BillToCode"])
        print(json.dumps("Output file generated with remark: 2"))

        outputFailedFile(outputFolderName, outputFileName, "Eoss_3", df_Eoss_3, "BillToCode", df_Summary["BillToCode"])
        print(json.dumps("Output file generated with remark: 3"))

        outputFailedFile(outputFolderName, outputFileName, "Promo_1", df_Promo_1, "BillToCode", df_Summary["BillToCode"])
        print(json.dumps("Output file generated with remark: 4"))

        outputFailedFile(outputFolderName, outputFileName, "Promo_2", df_Promo_2, "BillToCode", df_Summary["BillToCode"])
        print(json.dumps("Output file generated with remark: 5"))

        outputFailedFile(outputFolderName, outputFileName, "Promo_3", df_Promo_3, "BillToCode", df_Summary["BillToCode"])
        print(json.dumps("Output file generated with remark: 5.1"))

        file = open(os.path.join(outputFolderName, outputFileName), "rb")
        print(json.dumps("Output file generated with remark: 6"))

        if mode != 'dev':
            uploadFailedFile(inputFileName, file, clientUUID)
        print(json.dumps("Output file generated with remark: 7"))

        file.close()
        print(json.dumps("Output file generated with remark: 8"))
        deleteIfExists(os.path.join(outputFolderName, outputFileName))

    deleteIfExists(creditNoteNumberFilePath)
    deleteIfExists(partnerLocationsFilePath)
    print(json.dumps({"data":'Process complete', "code": "CMPLT", "isFailed" : len(df_Summary_Failed_List) > 0, 'isSuccess' : isSuccess}))
except Exception as e:
    deleteIfExists(creditNoteNumberFilePath)
    deleteIfExists(partnerLocationsFilePath)
    print(json.dumps({"data":"Error occured: " + str(e), "code": "ERROR"}))
